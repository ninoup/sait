from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form, Response, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer
from jose import jwt, JWTError
from datetime import datetime, timedelta
from typing import List, Optional
from pydantic import BaseModel
import os
import uuid
import openpyxl
from openpyxl import Workbook, load_workbook
from pathlib import Path
from sqlalchemy.orm import Session
from database import get_db, Student, Admin
from fastapi.staticfiles import StaticFiles
app = FastAPI(title="Olympiad Tracking System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

SECRET_KEY = "your-secret-key-here"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
EXCEL_FILE_PATH = "olympiads_data.xlsx"
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}
app.mount("/static", StaticFiles(directory="static"), name="static")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

if not Path(EXCEL_FILE_PATH).exists():
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Title", "Level", "Description", "Venue", "Date", 
               "Organizer", "Student ID", "Admin ID", "File Path"])
    wb.save(EXCEL_FILE_PATH)

class LoginRequest(BaseModel):
    username: str
    password: str

class Olympiad(BaseModel):
    id: int
    title: str
    level: str
    description: str
    venue: str
    date: str
    organizer: str
    student_id: int
    admin_id: int
    file_path: str

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_to_excel(olympiad_data: dict):
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    ws.append([
        olympiad_data["id"],
        olympiad_data["title"],
        olympiad_data["level"],
        olympiad_data["description"],
        olympiad_data["venue"],
        olympiad_data["date"],
        olympiad_data["organizer"],
        olympiad_data["student_id"],
        olympiad_data["admin_id"],
        olympiad_data["file_path"]
    ])
    wb.save(EXCEL_FILE_PATH)

def get_olympiads_from_excel() -> List[dict]:
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    olympiads = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        olympiad = {
            "id": row[0],
            "title": row[1],
            "level": row[2],
            "description": row[3],
            "venue": row[4],
            "date": row[5],
            "organizer": row[6],
            "student_id": row[7],
            "admin_id": row[8],
            "file_path": row[9]
        }
        olympiads.append(olympiad)
    
    return olympiads

def authenticate_user(username: str, password: str, db: Session):
    # Check if user is a student
    student = db.query(Student).filter(Student.username == username).first()
    if student and student.check_password(password):
        return {
            "user": student,
            "user_type": "student",
            "full_name": student.full_name,
            "id": student.id
        }
    
    # Check if user is an admin
    admin = db.query(Admin).filter(Admin.username == username).first()
    if admin and admin.check_password(password):
        return {
            "user": admin,
            "user_type": "admin",
            "full_name": admin.full_name,
            "id": admin.id
        }
    
    return None

@app.post("/login", response_model=dict)
async def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    auth_result = authenticate_user(login_data.username, login_data.password, db)
    if not auth_result:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(
        data={"sub": login_data.username, "user_type": auth_result["user_type"]}
    )
    
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "user_type": auth_result["user_type"],
        "full_name": auth_result["full_name"],
        "id": auth_result["id"]
    }

@app.get("/admins", response_model=List[dict])
async def get_admins(db: Session = Depends(get_db)):
    admins = db.query(Admin).all()
    return [{"id": admin.id, "full_name": admin.full_name} for admin in admins]

@app.get("/admin/olympiads", response_model=List[dict])
async def get_olympiads(db: Session = Depends(get_db)):
    try:
        olympiads = get_olympiads_from_excel()
        result = []
        
        for olympiad in olympiads:
            student = db.query(Student).filter(Student.id == olympiad["student_id"]).first()
            if not student:
                continue
            
            result.append({
                "id": olympiad["id"],
                "title": olympiad["title"],
                "level": olympiad["level"],
                "description": olympiad["description"],
                "venue": olympiad["venue"],
                "date": olympiad["date"],
                "organizer": olympiad["organizer"],
                "student_id": olympiad["student_id"],
                "admin_id": olympiad["admin_id"],
                "file_path": olympiad["file_path"],
                "student_full_name": student.full_name,
                "student_group": student.group
            })
        
        return result
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

@app.post("/olympiads", response_model=Olympiad)
async def create_olympiad(
    title: str = Form(...),
    level: str = Form(...),
    description: str = Form(...),
    venue: str = Form(...),
    date: str = Form(...),
    organizer: str = Form(...),
    student_id: int = Form(...),
    admin_id: int = Form(...),
    file: UploadFile = File(...)
):
    try:
        if not allowed_file(file.filename):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid file format. Allowed: pdf, png, jpg, jpeg, gif"
            )
        
        file_ext = file.filename.split('.')[-1]
        filename = f"{uuid.uuid4()}.{file_ext}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())
        
        olympiads = get_olympiads_from_excel()
        new_id = max([o["id"] for o in olympiads]) + 1 if olympiads else 1
        
        new_olympiad = {
            "id": new_id,
            "title": title,
            "level": level,
            "description": description,
            "venue": venue,
            "date": date,
            "organizer": organizer,
            "student_id": student_id,
            "admin_id": admin_id,
            "file_path": file_path
        }
        
        save_to_excel(new_olympiad)
        
        return new_olympiad
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

@app.get("/olympiads/{olympiad_id}/file")
async def get_olympiad_file(olympiad_id: int):
    try:
        olympiads = get_olympiads_from_excel()
        olympiad = next((o for o in olympiads if o["id"] == olympiad_id), None)
        
        if not olympiad:
            raise HTTPException(status_code=404, detail="Olympiad not found")
        
        file_path = olympiad["file_path"]
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        ext = file_path.split('.')[-1].lower()
        media_type = {
            'pdf': 'application/pdf',
            'png': 'image/png',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'gif': 'image/gif'
        }.get(ext, 'application/octet-stream')
        
        with open(file_path, "rb") as f:
            return Response(content=f.read(), media_type=media_type)
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)